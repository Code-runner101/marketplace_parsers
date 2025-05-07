from selenium.common import NoSuchElementException
from selenium import webdriver
from fake_useragent import UserAgent
import time
import random
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
import re
from selenium.webdriver.common.by import By
import openpyxl

ua = UserAgent()
options = uc.ChromeOptions()


# Функция для добавления задержки между запросами
def random_delay(min_delay=1, max_delay=15):
    time.sleep(random.uniform(min_delay, max_delay))


def get_value_by_param(table, param_name):
    # Ищем строку, содержащую нужный параметр
    row = table.find('span', string=param_name)
    if row:
        # Ищем родительский элемент <th> с параметром и получаем соседний элемент <td>
        parent_th = row.find_parent('th')
        if parent_th:
            # Ищем <td> в той же строке
            sibling_td = parent_th.find_next_sibling('td')
            if sibling_td:
                return sibling_td.get_text(strip=True)  # Возвращаем текст из <td>
    return "Не указано"  # Возвращаем пустую строку, если данные не найдены


def get_xpath(element):
    components = []
    child = element if element.name else element.parent
    for parent in child.parents:
        siblings = parent.find_all(child.name, recursive=False)
        if len(siblings) > 1:
            index = next(i for i, s in enumerate(siblings, 1) if s is child)
            components.append(f'{child.name}[{index}]')
        else:
            components.append(child.name)
        child = parent
    components.reverse()
    return '/{}'.format('/'.join(components))


'''WB посудомойки'''
def get_dwm_detales(product_url):
    # Случайный User-Agent
    random_user_agent = ua.random
    # Генерация случайного десктопного User-Agent
    while True:
        random_user_agent = ua.random
        if 'Mobile' not in random_user_agent and 'Android' not in random_user_agent and 'iPhone' not in random_user_agent:
            break  # Выход из цикла, если User-Agent десктопный
    options.add_argument(f"user-agent={random_user_agent}")
    driver = webdriver.Chrome(options=options)

    # Используем Selenium для перехода на страницу
    driver.get(product_url)

    # Ожидание загрузки страницы
    random_delay(2, 3)

    soup = BeautifulSoup(driver.page_source, 'html.parser')

    btn_texts = ["Все характеристики и описание", "Все характеристики"]

    # Переменные для хранения результатов
    dish_q_val = "Не указано"

    # Ищем элемент "Все характеристики" через BeautifulSoup
    for btn_text in btn_texts:
        specs_button = soup.find('button', class_="product-page__btn-detail hide-mobile", string=btn_text)
        if specs_button:
            break
    if specs_button:
        # Если найден, получаем XPATH его родительского элемента, чтобы кликнуть через Selenium
        specs_button_xpath = get_xpath(specs_button)
        try:
            button = driver.find_element(By.XPATH, specs_button_xpath)
            button.click()
            random_delay(2, 3)  # Ожидание появления всплывающего окна

            # Обновляем HTML после клика и появления окна с характеристиками
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            # Поиск всех таблиц
            table = soup.find('div', class_='product-params')

            # Проход по таблицам
            if table:
                dish_q_val_text = get_value_by_param(table, 'Вместимость комплектов посуды')
                if dish_q_val_text != "Не указано":
                    dish_q_val = re.sub(r'[^\d]', '', dish_q_val_text)

            return dish_q_val
        except NoSuchElementException:
            print("Элемент 'Все характеристики' не найден.")
            return dish_q_val
    else:
        return dish_q_val


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Посудомойки"
ws.append(['Название', 'Кол-во комплектов', 'Цена', 'Ссылка'])

# Загрузка HTML-контента из файла с основной страницы каталога
with open('/Users/macbookpro/Downloads/wb_dwm5.htm', 'r', encoding='utf-8') as file:
    soup = BeautifulSoup(file, 'html.parser')

    # Инициализируем счётчик строк
    row_counter = 2  # Начиная со второй строки, так как первая занята заголовками

    # Парсинг карточек товаров
    section = soup.find('div', class_='product-card-list')
    cards = section.find_all('div', class_='product-card__wrapper')

    for card in cards:
        # Извлекаем цену
        price = ""
        price_classes = [
            'price__lower-price wallet-price red-price',
            'price__lower-price',
            'price__lower-price wallet-price',
        ]

        for price_class in price_classes:
            price_tag = card.find('ins', class_=price_class)
            if price_tag:
                price = re.sub(r'[^\d]', '', price_tag.get_text(strip=True))
                break

        title_tag = card.find('span', class_='product-card__brand')
        title_text = card.find('span', class_='product-card__name')
        if title_tag and title_text:
            # Извлечение ссылки на товар
            title = title_tag.get_text(strip=True)
            title_text_val = title_text.get_text(strip=True)
            link_tag = card.find('a', class_='product-card__link j-card-link j-open-full-product-card')
            if link_tag:
                product_url = link_tag['href']
                if not product_url.startswith('https://'):
                    product_url = "https://www.wildberries.ru/" + product_url
                volume = get_dwm_detales(product_url)
                if not any(keyword in volume for keyword in ["6", "12", "14", "Не указано"]):
                    print("\n+---------------------------------------------------------+")
                    print(volume)
                    print(product_url)
                    print("+---------------------------------------------------------+\n")
                    continue
                if volume == "16":
                    continue
        else:
            title, product_link, volume, micro_power = "None"

        ws.append([title, volume, price])

        # Добавляем гиперссылку
        ws.cell(row=row_counter, column=4).hyperlink = product_url
        ws.cell(row=row_counter, column=4).value = "Перейти"
        ws.cell(row=row_counter, column=4).style = "Hyperlink"

        row_counter += 1
# Сохраняем файл
wb.save("/Users/macbookpro/Downloads/data_transfer_dwm.xlsx")