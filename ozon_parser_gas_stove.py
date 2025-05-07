from selenium.common import TimeoutException, NoSuchElementException
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from fake_useragent import UserAgent
import time
import random
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
import re
from selenium.webdriver.common.by import By
import openpyxl

ua = UserAgent()
options = uc.ChromeOptions()

# Функция для добавления задержки между запросами
def random_delay(min_delay=1, max_delay=15):
    time.sleep(random.uniform(min_delay, max_delay))


'''Парсер Ozon газ. плита'''
def get_oven_detales(product_url):
    # Случайный User-Agent
    random_user_agent = ua.random
    # Генерация случайного десктопного User-Agent
    while True:
        random_user_agent = ua.random
        if 'Mobile' not in random_user_agent and 'Android' not in random_user_agent and 'iPhone' not in random_user_agent:
            break  # Выход из цикла, если User-Agent десктопный
    options.add_argument(f"user-agent={random_user_agent}")
    driver = webdriver.Chrome(options=options)

    # Используем Selenium для перехода на страницу
    driver.get(product_url)

    # Ожидание загрузки страницы
    random_delay(2, 3)
    driver.refresh()
    random_delay(6, 7)

    # Получаем HTML-код после загрузки JavaScript
    soup = BeautifulSoup(driver.page_source, 'html.parser')

    # Переменные для хранения значений
    oven_type_val = "Не указано"
    oven_volume_text = "Не указано"
    burner_quant_val = "Не указано"

    features = soup.find('div', id='section-characteristics')
    if features:
        feature_info = soup.find_all('dl', class_='k3x_27')
        for feature in feature_info:
            # Извлекаем данные
            main_text_tag = feature.find('span', class_='x2k_27')
            page_text_tag = feature.find('dd', class_='xk2_27')
            if main_text_tag and page_text_tag:
                main_text = main_text_tag.get_text(strip=True)
                page_text = page_text_tag.get_text(strip=True)
                if 'Объем духовки, л' in main_text:
                    oven_volume_text = re.sub(r'\D', '', page_text)
                if 'Количество конфорок' in main_text:
                    burner_quant_val = re.sub(r'\D', '', page_text)
                if 'Тип духовки' in main_text:
                    oven_type_val = page_text

    return oven_type_val, oven_volume_text, burner_quant_val


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "газ-плита"
ws.append(['Название', 'Кол-во конфорок', 'Объем духовки', 'Тип духовки', 'Цена', 'Ссылка'])

# Список брендов для проверки и замены
brands = [
    "Gefest", "Gorenje", "NORDFROST", "Hansa", "KRAFT", "MAUNFELD", "Weissgauff", "Horizont", "Лысьва",
    "il Monte", "Beko", "MIU", "Simfer", "NEKО", "IDEAL", "Smeg", "Samsung", "DELVENTO", "Hurakan", "ЛАДА NOVA",
    "Bosch", "Leran", "Midea", "WILLMARK", "Candy", "Сармат", "Optima", "Haier", "Viomi", "ARTEL", "Гефест",
    "Apach", "Abat", "Aceline", "Bertazzoni", "Cezaris", "de luxe", "DARINA", "Darina", "De luxe",
    "Dauscher", "Danke", "DELUXE", "Deluxe", "Energy", "Flama", "Ferre", "Falken", "GEFEST",
    "GRILL MASTER", "Gemlux", "Hiberg", "Heimat", "INDOKOR", "itimat", "Kaiser", "KAYMAN",
    "Lexlight Shop", "MANYA", "M&G", "MQOUO", "Namilux", "RENOVA", "VESTA"
]

# Загрузка HTML-контента из файла с основной страницы каталога
with open('/Users/macbookpro/Downloads/ozon_gas3.htm', 'r', encoding='utf-8') as file:
    soup = BeautifulSoup(file, 'html.parser')

    # Инициализируем счётчик строк
    row_counter = 2  # Начиная со второй строки, так как первая занята заголовками

    # Парсинг карточек товаров
    cards = soup.find_all('div', class_='r1j_23 jr2_23')

    for card in cards:
        title_tag = card.find('a', class_='tile-hover-target oj7_23 jo8_23')
        if title_tag:
            # Извлечение ссылки на товар
            title = title_tag.get_text(strip=True)
            for brand in brands:
                if brand.lower() in title.lower():
                    title = brand  # Заменяем на название бренда
                    break
            product_url = title_tag['href']
            if not product_url.startswith('https://'):
                product_url = "https://www.mvideo.ru" + product_url

            oven_type, oven_volume, burner_quant = get_oven_detales(product_url)
        else:
            title, product_link, oven_type, oven_volume, burner_quant = "None"

        # Извлекаем цену
        # Поиск цены
        price = 'Не указано'
        price_div = card.find('div', class_='c3017-a0')
        if price_div:
            price_span = price_div.find('span', class_='c3017-a1 tsHeadline500Medium c3017-c0')
            if price_span:
                price_val = price_span.get_text(strip=True)
                price = re.sub(r'[^\d]', '', price_val)

        print(price)

        ws.append([title, burner_quant, oven_volume, oven_type, price])

        # Добавляем гиперссылку
        ws.cell(row=row_counter, column=6).hyperlink = product_url
        ws.cell(row=row_counter, column=6).value = "Перейти"
        ws.cell(row=row_counter, column=6).style = "Hyperlink"

        row_counter += 1
# Сохраняем файл
wb.save("/Users/macbookpro/Downloads/data_transfer_gas.xlsx")